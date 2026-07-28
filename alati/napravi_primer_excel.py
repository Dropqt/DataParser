#!/usr/bin/env python3
"""Napravi primer Excel fajla sa gotovom proverom JMBG-a.

    .venv/bin/python alati/napravi_primer_excel.py

Rezultat: ``primer/primer_gosti.xlsx``

Excel radi **istu proveru kao aplikacija** - kontrolnu cifru po zvaničnoj formuli - pa se
tipfeler vidi već u glavnoj tabeli, pre nego što se bilo šta kopira u program. Formule su
namerno pisane bez LET() i drugih novijih funkcija, da rade i u starijem Excelu i u
LibreOffice-u.

Raspored kolona je **isti kao izlaz iz aplikacije** (``models.EXPORT_HEADERS``), pa se
rezultat ture lepi preko A2 bez pomeranja ičega. Iza njih ide provera JMBG-a, pa pomoćne
kolone koje su sakrivene.

Slova kolona se nigde ne kucaju - izvode se iz ``HEADERS`` preko :func:`col`. Dodavanje
kolone je zato izmena na jednom mestu.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from eturista.validation import jmbg_check_digit  # noqa: E402

ROWS = 200  # dokle sežu formule i bojenje

#: Po jedan radni list za svaki nalog - svako popunjava svoje goste bez mešanja.
#: Nazivi se poklapaju sa ``ETURISTA_NALOGx_NAZIV`` iz ``.env``, pa se odmah vidi
#: koji list ide uz koji nalog u padajućem meniju aplikacije.
NALOZI = ["Danica", "Mileta", "Zorica"]

#: Primeri stoje zasebno da radni listovi budu prazni i spremni za pravi unos.
PRIMERI_SHEET = "Primeri"

#: Prve kolone moraju da budu istim redom kao EXPORT_HEADERS u aplikaciji, da bi se
#: rezultat ture lepio nazad preko A2 bez pomeranja ičega. Test to i čuva.
HEADERS = [
    ("Ime", 14),
    ("Prezime", 16),
    ("JMBG", 15),
    ("Dolazak", 12),
    ("Dana", 7),
    ("E-mail", 26),
    ("STATUS", 12),
    ("RAZLOG", 38),
    ("PDF", 26),
    ("PROVERA JMBG", 34),
]
#: Pomoćne kolone (sakrivene) - postoje samo da bi formula za proveru bila čitljiva.
#: ``sirov`` i ``čist`` rade isto što i ``validation.clean_jmbg``: sklone smeće koje dolazi
#: sa lepljenjem iz Worda i vrate vodeću nulu koju Excel pojede.
HELPERS = [
    ("sirov", 14), ("čist", 14),
    ("zbir", 8), ("kontrolna", 10), ("godina", 8), ("dat.rođ.", 10),
]

_NAZIVI = [naziv for naziv, _ in HEADERS + HELPERS]


def col(naziv: str) -> str:
    """Slovo kolone po naslovu iz zaglavlja.

    Slova se nigde ne kucaju ručno: dovoljno je dodati kolonu u ``HEADERS`` i sve
    formule, bojenja i opsezi se pomere sami. Ranije su bila ukucana na desetak mesta,
    pa je svaka nova kolona značila lov po celom fajlu.
    """
    return get_column_letter(_NAZIVI.index(naziv) + 1)


VISIBLE = "".join(col(naziv) for naziv, _ in HEADERS)
LAST_VISIBLE = col(HEADERS[-1][0])
FIRST_HELPER, LAST_HELPER = col(HELPERS[0][0]), col(HELPERS[-1][0])

#: Kolone koje korisnik popunjava - do njih seže bojenje reda po statusu.
LAST_INPUT = col("E-mail")

#: Koliko noćenja se podrazumeva. Mora da odgovara validation.DEFAULT_DAYS.
DEFAULT_DAYS = 5

GREEN = PatternFill("solid", start_color="FFC6EFCE")
RED = PatternFill("solid", start_color="FFFFC7CE")
YELLOW = PatternFill("solid", start_color="FFFFEB9C")
GRAY = PatternFill("solid", start_color="FFE0E0E0")
BLUE = PatternFill("solid", start_color="FFDDEBF7")
HEADER = PatternFill("solid", start_color="FF44546A")

GREEN_TEXT = Font(color="FF006100")
RED_TEXT = Font(color="FF9C0006")
YELLOW_TEXT = Font(color="FF9C6500")

THIN = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def jmbg(first_twelve: str) -> str:
    """Ispravan JMBG od prvih 12 cifara."""
    return first_twelve + str(jmbg_check_digit(first_twelve))


def broken_jmbg(first_twelve: str) -> str:
    """JMBG sa namerno pogrešnom kontrolnom cifrom - kao tipfeler u poslednjoj cifri.

    Cifra se pomera za jedan umesto da se upiše fiksna vrednost; fiksna bi za neke
    brojeve slučajno ispala tačna i primer ne bi pokazivao ništa.
    """
    correct = jmbg_check_digit(first_twelve)
    return first_twelve + str((correct + 1) % 10)


#: Izmišljeni primeri - ispravni po kontrolnoj cifri, ali ne pripadaju nikome.
#: Cifre su birane tako da izvedeni datum rođenja i pol odgovaraju imenu, jer se baš po
#: tom neslaganju najlakše primeti da je neko pogrešio cifru.
#: Poslednja dva reda su namerno pokvarena da se vidi kako provera reaguje.
#: Redom: ime, prezime, JMBG, datum.
#: Redom: ime, prezime, JMBG, datum dolaska, broj noćenja ("" = podrazumevanih 5).
SAMPLE = [
    #                            DD MM GGG RR BBB
    ("Marko", "Petrović", jmbg("010199071012"), "05.10.2026", ""),    # 01.01.1990, M
    ("Ana", "Jovanović", jmbg("230799575544"), "05.10.2026", ""),     # 23.07.1995, Ž
    ("Jovan", "Ilić", jmbg("150398871001"), "06.10.2026", 7),         # 15.03.1988, M
    ("Nikola", "Đorđević", jmbg("110599270012"), "06.10.2026", ""),   # 11.05.1992, M
    ("Milica", "Marković", jmbg("290200471505"), "07.10.2026", 10),   # 29.02.2004, Ž
    # pokvarena poslednja cifra -> "Pogrešna kontrolna cifra"
    ("Stefan", "Nikolić", broken_jmbg("070300771003"), "07.10.2026", ""),
    # 12 cifara, kao kad Excel pojede vodeću nulu -> žuto "⚠ Fali vodeća nula"
    ("Jelena", "Simić", jmbg("030697875501")[1:], "08.10.2026", ""),
]


#: Znaci koje lepljenje iz Worda ostavlja u ćeliji: razmak, tvrdi razmak, uski tvrdi
#: razmak, nevidljivi razmak nulte širine, crtica i crta na koju Word sam prepravi
#: crticu, pa tačka i kosa crta iz zapisa tipa ``010199-071012``. Aplikacija ionako
#: gleda samo cifre (``clean_jmbg``), pa i Excel treba isto.
#:
#: U formulu idu kao pravi znaci, ne preko ``CHAR(160)``: CHAR zavisi od kodne strane,
#: pa u LibreOffice-u daje drugi znak nego u Excelu i tvrdi razmak preživi čišćenje.
#: Ovde su pisani kao \u escape - u kodu mora da se vidi šta tačno stoji u formuli.
JUNK = (
    " ",        # običan razmak
    "\u00a0",   # tvrdi razmak (Ctrl+Shift+Space u Wordu)
    "\u202f",   # uski tvrdi razmak
    "\u200b",   # razmak nulte širine - stiže sa kopiranjem sa sajta
    "-",
    "\u2013",   # en crta, u nju Word sam prepravi crticu
    "\u2014",   # em crta
    ".",
    "/",
)


def ref(naziv: str) -> str:
    """Ćelija iz istog reda, ali tako da premeštanje reda ne može da je pokvari.

    Obična referenca (``C2``) je vezana za tačnu ćeliju, pa je Excel prepisuje čim se
    red iseče i prebaci na drugi list - a ako se posle prazan red obriše, ostaje
    ``#REF!``. To se dešava baš u najčešćoj situaciji: gost je upisan na pogrešan list,
    pa se seli kod drugog naloga.

    ``INDEX($C:$C,ROW())`` ne pokazuje ni na jednu određenu ćeliju nego znači "moja
    kolona, moj red" - preživljava i seljenje i brisanje redova, i na novom listu
    računa iz tog lista.
    """
    letter = col(naziv)
    return f"INDEX(${letter}:${letter},ROW())"


def build_formulas(row: int) -> dict[str, str]:
    """Formule za jedan red. ``row`` je broj reda u Excelu (podaci kreću od 2).

    Ključevi su ćelije u koje formula ide (tu red mora da se zna), a sve *reference*
    unutar formula idu preko :func:`ref` - videti tamo zašto.
    """
    raw, c = ref("JMBG"), ref("čist")
    sirov = ref("sirov")
    zbir, kontrolna = ref("zbir"), ref("kontrolna")
    godina, rodjen = ref("godina"), ref("dat.rođ.")

    def digit(position: int) -> str:
        return f"VALUE(MID({c},{position},1))"

    # Zvanična formula: 7(a+g) + 6(b+h) + 5(c+i) + 4(d+j) + 3(e+k) + 2(f+l)
    weighted = "+".join(
        f"{weight}*({digit(left)}+{digit(left + 6)})"
        for weight, left in zip((7, 6, 5, 4, 3, 2), range(1, 7))
    )

    year = f"IF(VALUE(MID({c},5,3))>=800,1000,2000)+VALUE(MID({c},5,3))"
    day, month = f"VALUE(MID({c},1,2))", f"VALUE(MID({c},3,2))"
    birth = f"DATE({godina},{month},{day})"

    # Ćelija kakva jeste, samo bez smeća: broj se ispisuje kao cifre (inače bi VALUE
    # radio nad "1,01991E+12"), a iz teksta se izbacuje ono što je Word doneo.
    cleaned = f"TRIM({raw})"
    for znak in JUNK:
        cleaned = f'SUBSTITUTE({cleaned},"{znak}","")'

    def cell(naziv: str) -> str:
        return f"{col(naziv)}{row}"

    return {
        # cifre kako su zaista upisane - po ovoj dužini se sudi šta fali
        cell("sirov"): f'=IF({raw}="","",IF(ISNUMBER({raw}),TEXT({raw},"0"),{cleaned}))',
        # 12 cifara znači da je pojedena vodeća nula (gost rođen 1-9. u mesecu);
        # vraćamo je i dalje računamo sa 13, isto kao clean_jmbg u aplikaciji
        cell("čist"): f'=IF({sirov}="","",IF(LEN({sirov})=12,"0"&{sirov},{sirov}))',
        # zbir sa težinama
        cell("zbir"): f'=IFERROR(IF(LEN({c})<>13,"",{weighted}),"")',
        # kontrolna cifra - ako ispadne 10 ili 11, ona je 0
        cell("kontrolna"): f'=IF({zbir}="","",IF(11-MOD({zbir},11)>9,0,11-MOD({zbir},11)))',
        # godina rođenja (GGG je godina po modulu 1000; 800+ znači 19xx)
        cell("godina"): f'=IFERROR(IF(LEN({c})<>13,"",{year}),"")',
        # datum rođenja - prazno ako taj datum ne postoji.
        # Excel DATE ne greši na 31.02, nego prevrne na mart, pa se dan i mesec proveravaju.
        cell("dat.rođ."): (
            f'=IFERROR(IF({godina}="","",'
            f'IF(AND(DAY({birth})={day},MONTH({birth})={month}),{birth},"")),"")'
        ),
        # broj noćenja se podrazumeva čim se upiše gost; prekucaj ako je drugačije
        cell("Dana"): f'=IF({ref("Ime")}="","",{DEFAULT_DAYS})',
        # poruka o ispravnosti JMBG-a - ista logika kao u aplikaciji.
        # Tri ishoda: ✓ ispravan, ⚠ radiće ali ćeliju treba popraviti, i sve ostalo (greška).
        # Sve je u IFERROR-u da bi i pokvarena formula rekla šta da se radi umesto #REF!.
        cell("PROVERA JMBG"): (
            f'=IFERROR(IF({raw}="","",'
            f'IF(LEN({c})<>13,"Nema 13 cifara (ima "&LEN({sirov})&")",'
            # kontrolna otpada na slovu među prvih 12 cifara; 13. se proverava posebno,
            # jer bi inače VALUE(MID(...,13,1)) niže pukao i ćelija bi pokazala #VALUE!
            f'IF(OR({kontrolna}="",ISERROR(VALUE(MID({c},13,1)))),'
            f'"JMBG sadrži nešto što nije cifra",'
            f'IF({rodjen}="","Nepostojeći datum rođenja",'
            f'IF(VALUE(MID({c},13,1))<>{kontrolna},'
            # Ako ni sa dodatom nulom kontrolna ne valja, onda cifra stvarno fali -
            # nije pojedena nula nego nedovršen unos.
            f'IF(LEN({sirov})=12,"Nema 13 cifara (ima 12)",'
            f'"Pogrešna kontrolna cifra - treba "&{kontrolna}),'
            f'IF(LEN({sirov})=12,"⚠ Fali vodeća nula - upiši 0"&{sirov},'
            f'IF(ISNUMBER({raw}),"⚠ Upisan kao broj - ćelija mora biti Tekst",'
            f'"✓ ispravan"))))))),'
            f'"Provera je pokvarena - prekopiraj ovu ćeliju iz ispravnog reda")'
        ),
    }


def build_guests_sheet(workbook: Workbook, title: str, with_samples: bool = False):
    """Jedan radni list sa gostima - zaglavlje, formule i bojenje.

    Pravi se po jedan za svaki nalog, pa svako popunjava svoje bez mešanja. Formule i
    bojenje su svuda isti; razlikuje se samo naziv lista i to da li nosi primere.
    """
    sheet = workbook.create_sheet(title)

    all_headers = HEADERS + HELPERS
    for index, (title, width) in enumerate(all_headers, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill = HEADER
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{LAST_VISIBLE}{ROWS + 1}"

    # Cela kolona kao Tekst, ne samo redova koliko ima formula: bez ovoga 201. gost
    # upada u ćeliju bez formata i Excel mu odmah pojede vodeću nulu.
    sheet.column_dimensions[col("JMBG")].number_format = "@"

    explicit_days: dict[int, int] = {}
    if with_samples:
        for offset, (given, surname, id_number, arrival, days) in enumerate(SAMPLE):
            row = offset + 2
            sheet[f"{col('Ime')}{row}"] = given
            sheet[f"{col('Prezime')}{row}"] = surname
            sheet[f"{col('JMBG')}{row}"] = id_number
            sheet[f"{col('Dolazak')}{row}"] = arrival
            if days:
                explicit_days[row] = int(days)

    for row in range(2, ROWS + 2):
        for reference, formula in build_formulas(row).items():
            sheet[reference] = formula

        # Gde je broj noćenja drugačiji od podrazumevanog, upisuje se preko formule -
        # isto što korisnik radi kad prekuca vrednost.
        if row in explicit_days:
            sheet[f"{col('Dana')}{row}"] = explicit_days[row]

        sheet[f"{col('JMBG')}{row}"].number_format = "@"   # tekst, da vodeća nula preživi
        sheet[f"{col('Dana')}{row}"].alignment = Alignment(horizontal="center")
        sheet[f"{col('STATUS')}{row}"].alignment = Alignment(horizontal="center")
        for column in VISIBLE:
            sheet[f"{column}{row}"].border = BOX

    # Pomoćne kolone su tu samo da bi formula za proveru bila čitljiva - sakrivamo ih.
    sheet.column_dimensions.group(FIRST_HELPER, LAST_HELPER, hidden=True)

    _add_conditional_formatting(sheet)
    return sheet


def _add_conditional_formatting(sheet) -> None:
    """Bojenje po ishodu provere i po statusu iz aplikacije.

    I ovde reference idu preko :func:`ref`, iz istog razloga kao u formulama: pravilo
    vezano za ``$J2`` se pri seljenju redova prepiše i ume da završi kao ``#REF!``, a
    tada bojenje tiho prestane da radi - crven red više ne pocrveni.
    """
    provera = col("PROVERA JMBG")
    span = f"{provera}2:{provera}{ROWS + 1}"
    sheet.conditional_formatting.add(
        span,
        FormulaRule(
            formula=[f'LEFT({ref("PROVERA JMBG")},1)="✓"'],
            fill=GREEN, font=GREEN_TEXT, stopIfTrue=True,
        ),
    )
    # Žuto: broj je ispravan i aplikacija bi ga primila, ali ćelija nije kako treba
    # (pojedena vodeća nula, upisano kao broj). Ide pre crvenog da ga ne pregazi.
    sheet.conditional_formatting.add(
        span,
        FormulaRule(
            formula=[f'LEFT({ref("PROVERA JMBG")},1)="⚠"'],
            fill=YELLOW, font=YELLOW_TEXT, stopIfTrue=True,
        ),
    )
    sheet.conditional_formatting.add(
        span,
        FormulaRule(formula=[f'{ref("PROVERA JMBG")}<>""'], fill=RED, font=RED_TEXT),
    )

    # JMBG koji je postao broj: lepljenje nosi format izvora i njime pregazi Tekst na
    # koloni. Boji se sama ćelija, a ne samo poruka desno - u poruku se ne gleda dok se
    # ne posumnja, a žuta ćelija usred kolone se vidi odmah. Ide pre bojenja reda po
    # statusu, da je zeleno "prijavljen" ne prekrije.
    jmbg = col("JMBG")
    sheet.conditional_formatting.add(
        f"{jmbg}2:{jmbg}{ROWS + 1}",
        FormulaRule(
            formula=[f'ISNUMBER({ref("JMBG")})'],
            fill=YELLOW, font=YELLOW_TEXT, stopIfTrue=True,
        ),
    )

    # STATUS kolona - popunjava se lepljenjem rezultata iz aplikacije.
    status_col = col("STATUS")
    status = f"{status_col}2:{status_col}{ROWS + 1}"
    for value, fill, font in (
        ("OK", GREEN, GREEN_TEXT),
        ("GREŠKA", RED, RED_TEXT),
        ("PRESKOČEN", GRAY, None),
        ("ČEKA", BLUE, None),
    ):
        rule = CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill)
        if font is not None:
            rule.font = font
        sheet.conditional_formatting.add(status, rule)

    # Red se blago oboji čim STATUS kaže da je gost pao - lakše se skenira lista.
    uneti = f"{col('Ime')}2:{LAST_INPUT}{ROWS + 1}"
    sheet.conditional_formatting.add(
        uneti,
        FormulaRule(formula=[f'{ref("STATUS")}="GREŠKA"'], fill=RED),
    )
    sheet.conditional_formatting.add(
        uneti,
        FormulaRule(formula=[f'{ref("STATUS")}="OK"'], fill=GREEN),
    )


def build_instructions_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Uputstvo")
    sheet.column_dimensions["A"].width = 4
    sheet.column_dimensions["B"].width = 110

    lines: list[tuple[str, str]] = [
        ("naslov", "Kako se koristi ova tabela"),
        ("", ""),
        ("podnaslov", "0. Listovi"),
        ("", f"Po jedan list za svaki nalog: {', '.join(NALOZI)}."),
        ("", "Svako popunjava svoj list, pa se grupe ne mešaju i svako vidi samo svoje goste."),
        ("", f"Nazivi listova su isti kao nalozi u padajućem meniju aplikacije."),
        ("", f"List '{PRIMERI_SHEET}' su izmišljeni primeri - tu se vidi kako radi provera"),
        ("", "JMBG-a. Radni listovi su namerno prazni."),
        ("", ""),
        ("podnaslov", "1. Unos gostiju"),
        ("", f"Popunjavaj kolone {col('Ime')}-{LAST_INPUT}: Ime, Prezime, JMBG, Dolazak, Dana, E-mail."),
        ("", "Cela JMBG kolona je formatirana kao TEKST - bez toga Excel pojede vodeću nulu"),
        ("", "kod gostiju rođenih 1-9. u mesecu, pa umesto 13 ostane 12 cifara."),
        ("", "Ako se nula ipak izgubi (lepljenje ume da pregazi format), kolona PROVERA"),
        ("", "požuti i napiše tačno šta treba da stoji u ćeliji. Sam unos i dalje prolazi -"),
        ("", "aplikacija tu nulu vraća sama - ali je bolje popraviti tabelu."),
        ("", "Dolazak piši kao 05.10 ili 05.10.2026 - oba se prepoznaju."),
        ("", f"Dana se popunjava samo od sebe na {DEFAULT_DAYS}; prekucaj ako je boravak duži."),
        ("", "E-mail popunjavaj samo za goste čiji vaučeri idu na drugu adresu nego ostali -"),
        ("", "prazna ćelija znači adresu upisanu u aplikaciji. Po njoj se prave folderi sa vaučerima."),
        ("", f"'{DEFAULT_DAYS} dana' znači {DEFAULT_DAYS} noćenja: dolazak 05.10 -> odlazak 10.10."),
        ("", ""),
        ("podnaslov", "2. Prebacivanje gosta na drugi list"),
        ("", "Kad se gost nađe kod pogrešnog naloga: označi njegove kolone A-F,"),
        ("", "Ctrl+C na drugi list, pa tek onda obriši original (Ctrl+X radi, ali ostavlja"),
        ("", "prazan red koji izgleda kao gost dok se ne obriše)."),
        ("", "Kolone G-P se ne prenose - one su formule i na svakom listu su iste."),
        ("", ""),
        ("podnaslov", "3. Kopiranje iz Worda"),
        ("", "Obično lepljenje (Ctrl+V) donosi i Wordov font, boje i okvire, pa tabela"),
        ("", "posle izgleda raštrkano i JMBG ume da se pretvori u broj bez vodeće nule."),
        ("", "Zato iz Worda uvek lepi BEZ FORMATA:"),
        ("", "  Excel: Ctrl+Shift+V, ili Početak > Nalepi > Zadrži samo tekst (Ctrl+Alt+V, pa T)"),
        ("", "  LibreOffice: Ctrl+Shift+V > Neformatirani tekst"),
        ("", "Tako ostaje format kolone (JMBG = Tekst) i bojenje ove tabele."),
        ("", ""),
        ("", "Još kraće: iz Worda može pravo u aplikaciju - označi u Wordu, pa u aplikaciji"),
        ("", "Ctrl+V. Aplikacija čita samo tekst, format je ne dodiruje, a polja prepoznaje"),
        ("", "sama. Ne mora ni da bude tabela: prolazi i običan spisak, jedan gost po redu"),
        ("", "('1. Marko Petrović 0101990710121 05.10.2026 5 dana')."),
        ("", ""),
        ("napomena", "Ako je formatiranje već upropašćeno: označi pokvarene redove, Početak >"),
        ("napomena", "Prenosilac formata sa nekog urednog reda, a JMBG koloni vrati format Tekst."),
        ("", ""),
        ("podnaslov", f"4. Provera JMBG-a (kolona {col('PROVERA JMBG')})"),
        ("", "Računa se automatski, po zvaničnoj formuli za kontrolnu cifru."),
        ("", "Zeleno '✓ ispravan' znači da broj matematički može da postoji."),
        ("", "Žuto '⚠' znači da je broj dobar ali ćelija nije - npr. fali vodeća nula."),
        ("", f"Ako požuti sama ćelija u koloni {col('JMBG')}, u nju je lepljenjem upao broj umesto"),
        ("", "teksta. Prekucaj je ili je nalepi ponovo bez formata - videti tačku 3."),
        ("", "Crveno kaže tačno šta ne valja, npr. 'Pogrešna kontrolna cifra - treba 8'."),
        ("", f"Kolone {FIRST_HELPER}-{LAST_HELPER} su pomoćne (sakrivene) - služe samo formuli u "
         f"koloni {col('PROVERA JMBG')}. Ne diraj ih."),
        ("", ""),
        ("napomena", "Provera hvata tipfelere, ali ne može da zna da li broj zaista pripada tom čoveku."),
        ("napomena", "To utvrđuje tek portal - i to se posle vidi u koloni RAZLOG."),
        ("", ""),
        ("podnaslov", "5. Slanje u aplikaciju"),
        ("", f"Označi kolone {col('Ime')}-{LAST_INPUT} za grupu gostiju koju prijavljuješ "
         "(5, 10, 30 - koliko hoćeš),"),
        ("", "Ctrl+C, pa u aplikaciji Ctrl+V."),
        ("", ""),
        ("podnaslov", "6. Vraćanje rezultata"),
        ("", f"Kad se tura završi, u aplikaciji Ctrl+C i ovde zalepi preko kolone {col('Ime')} za te goste."),
        ("", f"Popuniće se i {col('STATUS')} (STATUS), {col('RAZLOG')} (RAZLOG) i {col('PDF')} (PDF). "
         "Bojenje je već podešeno:"),
        ("", "OK = zeleno, GREŠKA = crveno, PRESKOČEN = sivo."),
        ("", f"Kolone {col('PROVERA JMBG')}-{LAST_HELPER} su formule i ostaju netaknute jer su "
         "desno od zalepljenog."),
        ("", ""),
        ("podnaslov", "Napomena"),
        ("", f"Na listu '{PRIMERI_SHEET}' su izmišljeni gosti - ne pripadaju nikome."),
        ("", "Poslednja dva su namerno pokvarena da se vidi kako provera reaguje:"),
        ("", "jednom je pokvarena kontrolna cifra, drugom fali vodeća nula."),
    ]

    for index, (kind, text) in enumerate(lines, start=1):
        cell = sheet.cell(row=index, column=2, value=text)
        if kind == "naslov":
            cell.font = Font(bold=True, size=15)
        elif kind == "podnaslov":
            cell.font = Font(bold=True, size=11, color="FF44546A")
        elif kind == "napomena":
            cell.font = Font(italic=True, color="FF9C0006")


def main() -> None:
    workbook = Workbook()
    # openpyxl novu radnu svesku pravi sa jednim praznim listom; svoje prave
    # ``create_sheet``, pa se taj podrazumevani sklanja.
    workbook.remove(workbook.active)

    for naziv in NALOZI:
        build_guests_sheet(workbook, naziv)
    build_guests_sheet(workbook, PRIMERI_SHEET, with_samples=True)
    build_instructions_sheet(workbook)

    workbook.active = 0   # fajl se otvara na prvom nalogu, ne na primerima

    target = Path(__file__).resolve().parent.parent / "primer" / "primer_gosti.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    print(f"Napravljeno: {target}")
    print(f"  listovi: {', '.join(NALOZI)} (prazni) + {PRIMERI_SHEET} ({len(SAMPLE)} primera) + Uputstvo")
    print(f"  formule do reda {ROWS + 1}")


if __name__ == "__main__":
    main()
